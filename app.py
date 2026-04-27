import csv
import io
import os
import re
from collections import Counter
from datetime import datetime

from flask import Flask, flash, make_response, redirect, render_template, request, session, url_for
from flask_sqlalchemy import SQLAlchemy
from reportlab.graphics import renderPDF
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'secretkey'

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/klasifikasi'
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
db = SQLAlchemy(app)

ALLOWED_UPLOAD_EXTENSIONS = {'csv', 'xlsx', 'xls'}
DATA_FIELDS = [
    'timestamp',
    'kelas',
    'nama_laoshi',
    'q1',
    'q2',
    'q3',
    'q4',
    'kritiksaran_laoshi',
    'kritiksaran_cetta',
]
FIELD_ALIASES = {
    'timestamp': {'timestamp', 'time', 'waktu', 'tanggal', 'datetime'},
    'kelas': {'kelas', 'class', 'label'},
    'nama_laoshi': {
        'nama_laoshi',
        'namalaoshi',
        'nama laoshi',
        'nama_pengajar',
        'namapengajar',
        'pengajar',
    },
    'q1': {
        'q1',
        'seberapa puas kamu dengan materi yang diberikan selama kelas?',
        'seberapa puas kamu dengan materi yang diberikan selama kelas',
    },
    'q2': {
        'q2',
        'seberapa paham kamu dengan materi yang diberikan selama kelas?',
        'seberapa paham kamu dengan materi yang diberikan selama kelas',
    },
    'q3': {
        'q3',
        'seberapa seru kegiatan belajar selama kelas?',
        'seberapa seru kegiatan belajar selama kelas',
    },
    'q4': {
        'q4',
        'menurut kamu, kecepatan laoshi dalam mengajar di kelas gimana?',
        'menurut kamu kecepatan laoshi dalam mengajar di kelas gimana?',
        'menurut kamu kecepatan laoshi dalam mengajar di kelas gimana',
    },
    'kritiksaran_laoshi': {
        'kritiksaran_laoshi',
        'kritik_saran_laoshi',
        'kritikdansaran_laoshi',
        'kritikdansaranlaoshi',
        'kritiksaran dosen',
        'kritik dan saran untuk laoshi kamu:',
        'kritik dan saran untuk laoshi kamu',
    },
    'kritiksaran_cetta': {
        'kritiksaran_cetta',
        'kritik_saran_cetta',
        'kritikdansaran_cetta',
        'kritikdansarancetta',
        'kritiksaran cetta',
        'kritik dan saran untuk cetta mandarin',
        'kritik dan saran untuk cetta mandarin:',
    },
}
TIMESTAMP_FORMATS = [
    '%d/%m/%Y %H:%M:%S',
    '%d/%m/%Y %H:%M',
    '%Y-%m-%d %H:%M:%S',
    '%Y-%m-%d %H:%M',
    '%Y-%m-%dT%H:%M',
]


def ensure_upload_folder():
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


def is_allowed_file(filename):
    return (
        '.' in filename
        and filename.rsplit('.', 1)[1].lower() in ALLOWED_UPLOAD_EXTENSIONS
    )


def normalize_key(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def get_recent_uploads(limit=10):
    ensure_upload_folder()
    uploaded_files = []

    for entry in os.scandir(app.config['UPLOAD_FOLDER']):
        if not entry.is_file():
            continue

        stats = entry.stat()
        uploaded_files.append(
            {
                'file': entry.name,
                'status': 'Uploaded',
                'tanggal': datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M'),
            }
        )

    uploaded_files.sort(key=lambda item: item['tanggal'], reverse=True)
    return uploaded_files[:limit]


def get_latest_uploaded_file():
    ensure_upload_folder()
    files = [entry for entry in os.scandir(app.config['UPLOAD_FOLDER']) if entry.is_file()]
    if not files:
        return None
    return max(files, key=lambda entry: entry.stat().st_mtime)


def normalize_headers(header_row, width):
    headers = []
    for index in range(width):
        raw_value = header_row[index] if index < len(header_row) else ''
        label = str(raw_value).strip() if raw_value is not None else ''
        headers.append(label or f'Kolom {index + 1}')
    return headers


def parse_csv_dataset(file_path):
    with open(file_path, newline='', encoding='utf-8-sig') as csv_file:
        rows = list(csv.reader(csv_file))
    return rows


def parse_xlsx_dataset(file_path):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_xls_dataset(file_path):
    import xlrd

    workbook = xlrd.open_workbook(file_path)
    worksheet = workbook.sheet_by_index(0)
    return [worksheet.row_values(index) for index in range(worksheet.nrows)]


def parse_dataset_rows(file_path):
    extension = file_path.rsplit('.', 1)[1].lower()
    if extension == 'csv':
        return parse_csv_dataset(file_path)
    if extension == 'xlsx':
        return parse_xlsx_dataset(file_path)
    if extension == 'xls':
        return parse_xls_dataset(file_path)
    raise ValueError('Format file tidak didukung.')


def clean_dataset_rows(raw_rows):
    cleaned_rows = []
    for row in raw_rows:
        normalized_row = ['' if cell is None else str(cell).strip() for cell in row]
        if any(value for value in normalized_row):
            cleaned_rows.append(normalized_row)
    return cleaned_rows


def get_uploaded_dataset_preview(limit=100):
    latest_file = get_latest_uploaded_file()
    if latest_file is None:
        return {'headers': [], 'rows': [], 'total_rows': 0, 'file_name': None, 'uploaded_at': None}

    raw_rows = parse_dataset_rows(latest_file.path)
    cleaned_rows = clean_dataset_rows(raw_rows)

    if not cleaned_rows:
        return {'headers': [], 'rows': [], 'total_rows': 0, 'file_name': latest_file.name, 'uploaded_at': None}

    headers = normalize_headers(cleaned_rows[0], max(len(row) for row in cleaned_rows))
    data_rows = []
    for row in cleaned_rows[1:limit + 1]:
        padded_row = row + [''] * (len(headers) - len(row))
        data_rows.append(padded_row[:len(headers)])

    return {
        'headers': headers,
        'rows': data_rows,
        'total_rows': max(len(cleaned_rows) - 1, 0),
        'file_name': latest_file.name,
        'uploaded_at': datetime.fromtimestamp(latest_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M'),
    }


def map_dataset_headers(headers):
    header_index = {}
    for idx, header in enumerate(headers):
        header_index[normalize_key(header)] = idx

    mapping = {}
    missing_fields = []
    for field in DATA_FIELDS:
        matched_index = None
        for alias in FIELD_ALIASES[field]:
            alias_index = header_index.get(normalize_key(alias))
            if alias_index is not None:
                matched_index = alias_index
                break

        if matched_index is None:
            missing_fields.append(field)
        else:
            mapping[field] = matched_index

    return mapping, missing_fields


def extract_structured_rows(file_path):
    raw_rows = parse_dataset_rows(file_path)
    cleaned_rows = clean_dataset_rows(raw_rows)
    if not cleaned_rows:
        raise ValueError('File tidak memiliki data yang bisa diproses.')

    headers = normalize_headers(cleaned_rows[0], max(len(row) for row in cleaned_rows))
    mapping, missing_fields = map_dataset_headers(headers)
    if missing_fields:
        missing = ', '.join(missing_fields)
        raise ValueError(f'Kolom wajib tidak ditemukan: {missing}.')

    records = []
    for row in cleaned_rows[1:]:
        padded_row = row + [''] * (len(headers) - len(row))
        record = {field: padded_row[mapping[field]] for field in DATA_FIELDS}
        if any(record.values()):
            records.append(record)

    if not records:
        raise ValueError('Baris data kosong. Tidak ada data yang disimpan.')

    return records


def build_batch_code(prefix):
    return f'{prefix}-{datetime.now().strftime("%Y%m%d%H%M%S")}'


def clear_all_batches():
    UploadBatch.query.delete()
    PreprocessingBatch.query.delete()
    ClassificationBatch.query.delete()
    HasilPreprocessing.query.delete()
    HasilKlasifikasi.query.delete()
    db.session.commit()


def clear_upload_folder():
    ensure_upload_folder()
    for entry in os.scandir(app.config['UPLOAD_FOLDER']):
        if entry.is_file():
            os.remove(entry.path)


def get_batch_rows(model, batch_code):
    return (
        model.query
        .filter(model.batch_code == batch_code)
        .order_by(model.id.asc())
        .all()
    )


def get_latest_batch_rows(model, batch_attr_name):
    latest_row = model.query.order_by(model.id.desc()).first()
    if latest_row is None:
        return None, []

    batch_value = getattr(latest_row, batch_attr_name)
    rows = (
        model.query
        .filter(getattr(model, batch_attr_name) == batch_value)
        .order_by(model.id.asc())
        .all()
    )
    return batch_value, rows


def get_latest_rows(model):
    latest_row = model.query.order_by(model.id.desc()).first()
    if latest_row is None:
        return []

    if hasattr(model, 'batch_code'):
        latest_batch = latest_row.batch_code
        return (
            model.query
            .filter(model.batch_code == latest_batch)
            .order_by(model.id.asc())
            .all()
        )

    latest_time = latest_row.created_at
    return (
        model.query
        .filter(model.created_at == latest_time)
        .order_by(model.id.asc())
        .all()
    )


def build_chart_data(classification_rows):
    counts = Counter((row.hasil_klasifikasi or 'Belum ada hasil') for row in classification_rows)
    return {
        'labels': list(counts.keys()),
        'values': list(counts.values()),
    }


def build_rekapan_summary(classification_rows):
    counts = Counter((row.hasil_klasifikasi or 'Belum ada hasil') for row in classification_rows)
    cards = []
    accents = ['success', 'primary', 'warning', 'secondary']
    for index, (label, total) in enumerate(counts.items()):
        cards.append(
            {
                'title': label,
                'value': total,
                'accent': accents[index % len(accents)],
            }
        )
    return cards


def parse_row_timestamp(value):
    raw_value = str(value or '').strip()
    if not raw_value:
        return None

    for fmt in TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(raw_value, fmt)
        except ValueError:
            continue
    return None


def filter_rows_by_timestamp_range(rows, start_value=None, end_value=None):
    start_dt = parse_row_timestamp(start_value) if start_value else None
    end_dt = parse_row_timestamp(end_value) if end_value else None
    if end_dt is not None and end_value and len(end_value.strip()) == 16:
        end_dt = end_dt.replace(second=59)

    filtered_rows = []
    for row in rows:
        row_dt = parse_row_timestamp(row.timestamp)
        if row_dt is None:
            continue
        if start_dt and row_dt < start_dt:
            continue
        if end_dt and row_dt > end_dt:
            continue
        filtered_rows.append(row)
    return filtered_rows


def build_pdf_response(title, subtitle, classification_rows, filename):
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles['Title']),
        Spacer(1, 10),
    ]

    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
        elements.append(Spacer(1, 12))

    chart_data = build_chart_data(classification_rows)
    if chart_data['labels'] and chart_data['values']:
        drawing = Drawing(360, 220)
        pie = Pie()
        pie.x = 100
        pie.y = 20
        pie.width = 160
        pie.height = 160
        pie.data = chart_data['values']
        pie.labels = [f'{label} ({value})' for label, value in zip(chart_data['labels'], chart_data['values'])]
        pie.slices.strokeWidth = 0.5
        pie.slices[0].fillColor = colors.HexColor('#1f6feb')
        palette = ['#1f6feb', '#2da44e', '#f59e0b', '#ef4444', '#8b5cf6', '#06b6d4']
        for index, color in enumerate(palette[:len(chart_data['values'])]):
            pie.slices[index].fillColor = colors.HexColor(color)
        drawing.add(String(0, 205, 'Pie Chart Hasil Klasifikasi', fontSize=12, fillColor=colors.HexColor('#0f172a')))
        drawing.add(pie)
        elements.append(drawing)
        elements.append(Spacer(1, 12))

    table_rows = [[
        'No',
        'Timestamp',
        'Nama File',
        'Kelas',
        'Nama Laoshi',
        'Q1',
        'Q2',
        'Q3',
        'Q4',
        'Hasil Klasifikasi',
    ]]
    for index, row in enumerate(classification_rows, start=1):
        table_rows.append([
            str(index),
            row.timestamp,
            row.source_file,
            row.kelas,
            row.nama_laoshi,
            row.q1 or '',
            row.q2 or '',
            row.q3 or '',
            row.q4 or '',
            row.hasil_klasifikasi,
        ])

    pdf_table = Table(table_rows, repeatRows=1)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f6feb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(pdf_table)

    document.build(elements)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def get_saved_classification_rows():
    return (
        HasilKlasifikasi.query
        .filter_by(is_saved=True)
        .order_by(HasilKlasifikasi.saved_at.asc(), HasilKlasifikasi.id.asc())
        .all()
    )


def get_upload_batches():
    return UploadBatch.query.order_by(UploadBatch.created_at.desc(), UploadBatch.id.desc()).all()


def get_preprocessing_batches():
    return PreprocessingBatch.query.order_by(PreprocessingBatch.created_at.desc(), PreprocessingBatch.id.desc()).all()


def get_classification_batches():
    return ClassificationBatch.query.order_by(ClassificationBatch.created_at.desc(), ClassificationBatch.id.desc()).all()


def get_upload_batch_preview(upload_batch):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], upload_batch.saved_name)
    raw_rows = parse_dataset_rows(file_path)
    cleaned_rows = clean_dataset_rows(raw_rows)
    if not cleaned_rows:
        return {'headers': [], 'rows': []}

    headers = normalize_headers(cleaned_rows[0], max(len(row) for row in cleaned_rows))
    rows = []
    for row in cleaned_rows[1:51]:
        padded_row = row + [''] * (len(headers) - len(row))
        rows.append(padded_row[:len(headers)])
    return {'headers': headers, 'rows': rows}


def get_preprocessing_batch_preview(pre_batch):
    rows = get_batch_rows(HasilPreprocessing, pre_batch.batch_code)
    headers = ['Timestamp', 'Kelas', 'Nama Laoshi', 'Q1', 'Q2', 'Q3', 'Q4', 'Kritik Saran Laoshi', 'Kritik Saran Cetta', 'Hasil Preprocessing']
    values = [
        [row.timestamp, row.kelas, row.nama_laoshi, row.q1, row.q2, row.q3, row.q4, row.kritiksaran_laoshi, row.kritiksaran_cetta, row.hasil_preprocessing]
        for row in rows[:50]
    ]
    return {'headers': headers, 'rows': values}


def get_classification_batch_preview(cls_batch):
    rows = get_batch_rows(HasilKlasifikasi, cls_batch.batch_code)
    headers = ['Timestamp', 'Kelas', 'Nama Laoshi', 'Q1', 'Q2', 'Q3', 'Q4', 'Hasil Preprocessing', 'Hasil Klasifikasi']
    values = [
        [row.timestamp, row.kelas, row.nama_laoshi, row.q1, row.q2, row.q3, row.q4, row.hasil_preprocessing, row.hasil_klasifikasi]
        for row in rows[:50]
    ]
    return {'headers': headers, 'rows': values}


def is_ajax_request():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def render_page(template_name, partial_name, **context):
    if is_ajax_request():
        return render_template(partial_name, **context)
    return render_template(template_name, **context)


class users(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True, nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)


class DataPengajar(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama_pengajar = db.Column(db.String(250), unique=True, nullable=False)


class UploadBatch(db.Model):
    __tablename__ = 'upload_batch'

    id = db.Column(db.Integer, primary_key=True)
    batch_code = db.Column(db.String(40), nullable=False, unique=True, index=True)
    file_name = db.Column(db.String(255), nullable=False)
    saved_name = db.Column(db.String(255), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


class PreprocessingBatch(db.Model):
    __tablename__ = 'preprocessing_batch'

    id = db.Column(db.Integer, primary_key=True)
    batch_code = db.Column(db.String(40), nullable=False, unique=True, index=True)
    upload_batch_id = db.Column(db.Integer, nullable=False, index=True)
    file_name = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


class ClassificationBatch(db.Model):
    __tablename__ = 'classification_batch'

    id = db.Column(db.Integer, primary_key=True)
    batch_code = db.Column(db.String(40), nullable=False, unique=True, index=True)
    preprocessing_batch_id = db.Column(db.Integer, nullable=False, index=True)
    file_name = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


class HasilPreprocessing(db.Model):
    __tablename__ = 'hasil_preprocessing'

    id = db.Column(db.Integer, primary_key=True)
    batch_code = db.Column(db.String(40), nullable=False, index=True)
    source_file = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    timestamp = db.Column(db.String(100), nullable=False)
    kelas = db.Column(db.String(100), nullable=False)
    nama_laoshi = db.Column(db.String(255), nullable=False)
    q1 = db.Column(db.String(50), nullable=True)
    q2 = db.Column(db.String(50), nullable=True)
    q3 = db.Column(db.String(50), nullable=True)
    q4 = db.Column(db.String(50), nullable=True)
    kritiksaran_laoshi = db.Column(db.Text, nullable=True)
    kritiksaran_cetta = db.Column(db.Text, nullable=True)
    hasil_preprocessing = db.Column(db.Text, nullable=False)


class HasilKlasifikasi(db.Model):
    __tablename__ = 'hasil_klasifikasi'

    id = db.Column(db.Integer, primary_key=True)
    batch_code = db.Column(db.String(40), nullable=False, index=True)
    preprocessing_batch = db.Column(db.String(40), nullable=False, index=True)
    source_file = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    saved_at = db.Column(db.DateTime, nullable=True)
    is_saved = db.Column(db.Boolean, nullable=False, default=False)
    timestamp = db.Column(db.String(100), nullable=False)
    kelas = db.Column(db.String(100), nullable=False)
    nama_laoshi = db.Column(db.String(255), nullable=False)
    q1 = db.Column(db.String(50), nullable=True)
    q2 = db.Column(db.String(50), nullable=True)
    q3 = db.Column(db.String(50), nullable=True)
    q4 = db.Column(db.String(50), nullable=True)
    kritiksaran_laoshi = db.Column(db.Text, nullable=True)
    kritiksaran_cetta = db.Column(db.Text, nullable=True)
    hasil_preprocessing = db.Column(db.Text, nullable=False)
    hasil_klasifikasi = db.Column(db.String(100), nullable=False)


with app.app_context():
    db.create_all()


@app.route('/')
def home():
    return redirect('/login')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user_data = users.query.filter_by(username=username).first()

        if user_data and user_data.password == password:
            session['user'] = user_data.username
            return redirect(url_for('dashboard'))

        flash('Username atau Password salah!', 'danger')
        return redirect(url_for('login'))

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        new_user = users(
            username=username,
            email=email,
            password=password
        )

        db.session.add(new_user)
        db.session.commit()

        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect('/login')

    username_login = session['user']
    recent_teachers = DataPengajar.query.order_by(DataPengajar.id.desc()).limit(5).all()
    upload_batches = get_upload_batches()
    preprocessing_batches = get_preprocessing_batches()
    classification_batches = get_classification_batches()

    stats = [
        {"title": "Total Pengajar", "value": DataPengajar.query.count(), "icon": "fa-users", "accent": "primary"},
        {"title": "Upload Batch", "value": len(upload_batches), "icon": "fa-upload", "accent": "secondary"},
        {"title": "Preprocess Batch", "value": len(preprocessing_batches), "icon": "fa-file-text-o", "accent": "success"},
        {"title": "Klasifikasi Batch", "value": len(classification_batches), "icon": "fa-refresh", "accent": "warning"},
    ]

    return render_page(
        'index.html',
        '_dashboard_content.html',
        nama_user=username_login,
        active_page='dashboard',
        page_title='Dashboard',
        page_subtitle='Ringkasan sistem klasifikasi dan akses cepat ke tiap modul.',
        stats=stats,
        recent_teachers=recent_teachers
    )


@app.route('/data_pengajar')
def data_pengajar():
    if 'user' not in session:
        return redirect(url_for('login'))

    search_query = request.args.get('q', '', type=str).strip()
    page = request.args.get('page', 1, type=int)

    query = DataPengajar.query.order_by(DataPengajar.nama_pengajar.asc())
    if search_query:
        query = query.filter(DataPengajar.nama_pengajar.ilike(f'%{search_query}%'))

    pengajar_paginated = query.paginate(page=page, per_page=10, error_out=False)

    return render_page(
        'data_pengajar.html',
        '_data_pengajar_content.html',
        data=pengajar_paginated,
        search_query=search_query,
        active_page='data_pengajar',
        page_title='Data Pengajar',
        page_subtitle='Kelola data pengajar dan cari entri yang sudah tersimpan.'
    )


@app.route('/rekapan')
def rekapan():
    if 'user' not in session:
        return redirect(url_for('login'))

    classification_rows = get_saved_classification_rows()
    summary_cards = build_rekapan_summary(classification_rows)
    classification_chart = build_chart_data(classification_rows)
    classification_batches = get_classification_batches()

    classification_batch_rows = [
        {
            'no': index,
            'id': batch.id,
            'timestamp': batch.created_at.strftime('%Y-%m-%d %H:%M'),
            'file_name': batch.file_name,
        }
        for index, batch in enumerate(classification_batches, start=1)
    ]

    return render_page(
        'rekapan.html',
        '_rekapan_content.html',
        summary_cards=summary_cards,
        classification_batches=classification_batch_rows,
        classification_chart=classification_chart,
        active_page='rekapan',
        page_title='Rekapan Klasifikasi',
        page_subtitle='Ringkasan hasil klasifikasi dan performa tiap pengajar.'
    )


@app.route('/preprocessing')
def preprocessing_redirect():
    return redirect(url_for('input_data'))


@app.route('/input_data', methods=['GET', 'POST'])
def input_data():
    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        if action == 'upload':
            uploaded_file = request.files.get('dataset_file')

            if uploaded_file is None or uploaded_file.filename == '':
                flash('Pilih file terlebih dahulu sebelum upload.', 'danger')
                return redirect(url_for('input_data'))

            if not is_allowed_file(uploaded_file.filename):
                flash('Format file tidak didukung. Gunakan CSV, XLSX, atau XLS.', 'danger')
                return redirect(url_for('input_data'))

            ensure_upload_folder()
            original_name = secure_filename(uploaded_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            saved_name = f'{timestamp}_{original_name}'
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_name)

            uploaded_file.save(save_path)
            upload_batch = UploadBatch(
                batch_code=build_batch_code('UPL'),
                file_name=original_name,
                saved_name=saved_name,
            )
            db.session.add(upload_batch)
            db.session.commit()
            flash(f'File {original_name} berhasil diupload.', 'success')
            return redirect(url_for('input_data', tab='upload'))

        if action == 'preprocess':
            upload_batch_id = request.form.get('upload_batch_id', type=int)
            upload_batch = UploadBatch.query.get(upload_batch_id)
            if upload_batch is None:
                flash('Batch upload tidak ditemukan.', 'danger')
                return redirect(url_for('input_data', tab='upload'))

            existing_pre_batch = PreprocessingBatch.query.filter_by(upload_batch_id=upload_batch.id).first()
            if existing_pre_batch is not None:
                cls_batches = ClassificationBatch.query.filter_by(preprocessing_batch_id=existing_pre_batch.id).all()
                for cls_batch in cls_batches:
                    HasilKlasifikasi.query.filter_by(batch_code=cls_batch.batch_code).delete()
                    db.session.delete(cls_batch)
                HasilPreprocessing.query.filter_by(batch_code=existing_pre_batch.batch_code).delete()
                db.session.delete(existing_pre_batch)
                db.session.commit()

            file_path = os.path.join(app.config['UPLOAD_FOLDER'], upload_batch.saved_name)
            try:
                records = extract_structured_rows(file_path)
            except ValueError as error:
                flash(str(error), 'danger')
                return redirect(url_for('input_data', tab='upload'))

            batch_code = build_batch_code('PRE')
            preprocessing_batch = PreprocessingBatch(
                batch_code=batch_code,
                upload_batch_id=upload_batch.id,
                file_name=upload_batch.file_name,
            )
            db.session.add(preprocessing_batch)
            rows = []
            for record in records:
                rows.append(
                    HasilPreprocessing(
                        batch_code=batch_code,
                        source_file=upload_batch.file_name,
                        timestamp=record['timestamp'],
                        kelas=record['kelas'],
                        nama_laoshi=record['nama_laoshi'],
                        q1=record['q1'],
                        q2=record['q2'],
                        q3=record['q3'],
                        q4=record['q4'],
                        kritiksaran_laoshi=record['kritiksaran_laoshi'],
                        kritiksaran_cetta=record['kritiksaran_cetta'],
                        hasil_preprocessing='Data siap untuk tahap preprocessing lanjutan',
                    )
                )

            db.session.add_all(rows)
            db.session.commit()
            flash(f'Batch {upload_batch.file_name} berhasil diproses.', 'success')
            return redirect(url_for('input_data', tab='preprocessing'))

        if action == 'classify':
            preprocessing_batch_id = request.form.get('preprocessing_batch_id', type=int)
            preprocessing_batch = PreprocessingBatch.query.get(preprocessing_batch_id)
            if preprocessing_batch is None:
                flash('Batch preprocessing tidak ditemukan.', 'danger')
                return redirect(url_for('input_data', tab='preprocessing'))

            preprocessing_rows = get_batch_rows(HasilPreprocessing, preprocessing_batch.batch_code)
            if not preprocessing_rows:
                flash('Belum ada data preprocessing yang bisa diklasifikasikan.', 'danger')
                return redirect(url_for('input_data', tab='preprocessing'))

            existing_cls_batch = ClassificationBatch.query.filter_by(preprocessing_batch_id=preprocessing_batch.id).first()
            if existing_cls_batch is not None:
                HasilKlasifikasi.query.filter_by(batch_code=existing_cls_batch.batch_code).delete()
                db.session.delete(existing_cls_batch)
                db.session.commit()

            batch_code = build_batch_code('CLS')
            classification_batch = ClassificationBatch(
                batch_code=batch_code,
                preprocessing_batch_id=preprocessing_batch.id,
                file_name=preprocessing_batch.file_name,
            )
            db.session.add(classification_batch)
            classification_rows = []
            for row in preprocessing_rows:
                classification_rows.append(
                    HasilKlasifikasi(
                        batch_code=batch_code,
                        preprocessing_batch=preprocessing_batch.batch_code,
                        source_file=row.source_file,
                        saved_at=datetime.utcnow(),
                        is_saved=True,
                        timestamp=row.timestamp,
                        kelas=row.kelas,
                        nama_laoshi=row.nama_laoshi,
                        q1=row.q1,
                        q2=row.q2,
                        q3=row.q3,
                        q4=row.q4,
                        kritiksaran_laoshi=row.kritiksaran_laoshi,
                        kritiksaran_cetta=row.kritiksaran_cetta,
                        hasil_preprocessing=row.hasil_preprocessing,
                        hasil_klasifikasi=row.kelas or 'Belum ditentukan',
                    )
                )

            db.session.add_all(classification_rows)
            db.session.commit()
            flash(f'Batch {preprocessing_batch.file_name} berhasil diklasifikasikan.', 'success')
            return redirect(url_for('input_data', tab='classification'))

    active_tab = request.args.get('tab', 'upload')

    if active_tab not in {'upload', 'preprocessing', 'classification'}:
        active_tab = 'upload'

    return render_page(
        'preprocessing.html',
        '_preprocessing_content.html',
        upload_batches=get_upload_batches(),
        preprocessing_batches=get_preprocessing_batches(),
        classification_batches=get_classification_batches(),
        active_tab=active_tab,
        active_page='input_data',
        page_title='Input Data',
        page_subtitle='Kelola batch upload, preprocessing, dan klasifikasi secara bertahap.'
    )


@app.route('/input_data/download_classification')
def download_classification():
    if 'user' not in session:
        return redirect(url_for('login'))

    classification_batch_id = request.args.get('classification_batch_id', type=int)
    classification_batch = ClassificationBatch.query.get(classification_batch_id)
    if classification_batch is None:
        flash('Batch klasifikasi tidak ditemukan.', 'danger')
        return redirect(url_for('input_data', tab='classification'))

    classification_rows = get_batch_rows(HasilKlasifikasi, classification_batch.batch_code)
    if not classification_rows:
        flash('Belum ada hasil klasifikasi untuk didownload.', 'danger')
        return redirect(url_for('input_data', tab='classification'))

    return build_pdf_response(
        title='Hasil Klasifikasi Batch',
        subtitle=f'Batch: {classification_batch.file_name} | Dibuat: {classification_batch.created_at.strftime("%Y-%m-%d %H:%M")}',
        classification_rows=classification_rows,
        filename=f'hasil_klasifikasi_{classification_batch.batch_code}.pdf',
    )


@app.route('/input_data/batch_preview/<string:batch_type>/<int:batch_id>')
def batch_preview(batch_type, batch_id):
    if 'user' not in session:
        return redirect(url_for('login'))

    preview_title = ''
    preview_data = {'headers': [], 'rows': []}

    if batch_type == 'upload':
        batch = UploadBatch.query.get(batch_id)
        if batch is None:
            return '<p class="text-muted mb-0">Batch upload tidak ditemukan.</p>'
        preview_title = f'Preview Upload: {batch.file_name}'
        preview_data = get_upload_batch_preview(batch)
    elif batch_type == 'preprocessing':
        batch = PreprocessingBatch.query.get(batch_id)
        if batch is None:
            return '<p class="text-muted mb-0">Batch preprocessing tidak ditemukan.</p>'
        preview_title = f'Preview Preprocessing: {batch.file_name}'
        preview_data = get_preprocessing_batch_preview(batch)
    elif batch_type == 'classification':
        batch = ClassificationBatch.query.get(batch_id)
        if batch is None:
            return '<p class="text-muted mb-0">Batch klasifikasi tidak ditemukan.</p>'
        preview_title = f'Preview Klasifikasi: {batch.file_name}'
        preview_data = get_classification_batch_preview(batch)
    else:
        return '<p class="text-muted mb-0">Tipe batch tidak dikenal.</p>'

    return render_template(
        '_batch_preview_modal_content.html',
        preview_title=preview_title,
        preview_headers=preview_data['headers'],
        preview_rows=preview_data['rows'],
    )


@app.route('/rekapan/download')
def download_rekapan_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    start_timestamp = request.args.get('start_timestamp', '', type=str).strip()
    end_timestamp = request.args.get('end_timestamp', '', type=str).strip()
    classification_rows = get_saved_classification_rows()
    filtered_rows = filter_rows_by_timestamp_range(classification_rows, start_timestamp, end_timestamp)

    if not filtered_rows:
        flash('Tidak ada data klasifikasi pada rentang timestamp tersebut.', 'danger')
        return redirect(url_for('rekapan'))

    subtitle_parts = []
    if start_timestamp:
        subtitle_parts.append(f'Mulai: {start_timestamp}')
    if end_timestamp:
        subtitle_parts.append(f'Sampai: {end_timestamp}')
    subtitle = ' | '.join(subtitle_parts) if subtitle_parts else 'Semua data klasifikasi tersimpan'

    return build_pdf_response(
        title='Rekapan Klasifikasi',
        subtitle=subtitle,
        classification_rows=filtered_rows,
        filename='rekapan_klasifikasi.pdf',
    )


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


if __name__ == "__main__":
    app.run(debug=True)
